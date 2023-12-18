# -*- coding: utf-8 -*-
"""
Created on Fri Jul 21 13:32:36 2023

@author: johan
"""

import os

def print_directory_structure(path, indentation_level=0):
    """
    Prints the directory structure starting from the given path.
    
    Args:
        path (str): The path to the directory structure to print.
        indentation_level (int): The current level of indentation.
    """
    # Print the current directory
    print('\t' * indentation_level + os.path.basename(path))
    
    # If the current path is a directory, recursively print its contents
    if os.path.isdir(path):
        for child_path in os.listdir(path):
            print_directory_structure(os.path.join(path, child_path), indentation_level + 1)

# Example of use
print_directory_structure("C:\\Users\\johan\\Bureau\\Data_Lineage")


import pandas as pd
import os
from openpyxl import load_workbook

def extract_excel_info(directory):
    # Dict to hold file : sheet : column names mapping
    file_sheet_column_names = {}

    # Walk through the directory
    for root, dirs, files in os.walk(directory):
        for file in files:
            # Check if the file is an Excel file
            if file.endswith('.xlsx'):
                # Full file path
                file_path = os.path.join(root, file)
                
                # Load the workbook
                workbook = load_workbook(filename=file_path)

                # Get sheet names
                sheet_names = workbook.sheetnames

                # Dict to hold sheet : column names mapping
                sheet_column_names = {}

                for sheet in sheet_names:
                    # Read the sheet into a pandas DataFrame
                    df = pd.read_excel(file_path, sheet_name=sheet)

                    # Get the column names
                    column_names = df.columns.tolist()

                    # Add to dict
                    sheet_column_names[sheet] = column_names
                
                # Add to overall dict
                file_sheet_column_names[file] = sheet_column_names

    return file_sheet_column_names


extract_excel_info('C:\\Users\\johan\\Bureau\\Data_Lineage\\data_lineage_project\\data\\Insurance_Company')
