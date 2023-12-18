# -*- coding: utf-8 -*-
"""
Created on Fri Jul 21 15:32:04 2023

@author: johan
"""

import os
import pytest
import pandas as pd
from openpyxl import load_workbook
from ..src import main

@pytest.fixture
def run_main_and_get_files():
    # Exécuter le script principal
    main()
    
    # Les noms des fichiers générés
    files = ['Pricing.xlsx', 'Reserving.xlsx', 'RiskManagement.xlsx', 'Solvency.xlsx', 'Capital_Modeling.xlsx']
    
    # Les répertoires des fichiers générés
    directories = ['Pricing_directory', 'Reserving_directory', 'RiskManagement_directory', 'Solvency_directory', 'Capital_Modeling_directory']
    
    # Renvoyer les chemins complets des fichiers générés
    paths = [os.path.join(directory, file) for file, directory in zip(files, directories)]
    return paths


def test_main(run_main_and_get_files):
    # Vérifier si les fichiers ont bien été créés
    for file in run_main_and_get_files:
        assert os.path.isfile(file), f"Le fichier {file} n'a pas été créé."
    
    # Pour chaque fichier, charger le workbook et vérifier si les onglets attendus sont présents
    for file in run_main_and_get_files:
        wb = load_workbook(file)
        sheets = wb.sheetnames
        
        # Vérifier si les onglets attendus sont présents
        if 'Pricing' in file:
            expected_sheets = ['Base_Params', 'Demographic_Data', 'Claim_Data', 'Premium_Data', 'Actuarial_Calculations', 'Additional_Calculations_6', 'Additional_Calculations_7', 'Additional_Calculations_8', 'Additional_Calculations_9', 'Additional_Calculations_10', 'Additional_Calculations_11', 'Additional_Calculations_12']
            assert all(sheet in sheets for sheet in expected_sheets), f"Il manque des onglets dans le fichier {file}."
        
        # Répétez le processus similaire pour les autres fichiers.
        # Veuillez noter que les noms des onglets attendus doivent être modifiés en fonction de vos attentes exactes pour chaque fichier.
