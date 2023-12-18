# import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from generation.data_generator import DataGenerator
from generation.excel_generator import ExcelGenerator
from graph_generator import GraphGenerator

import os
import shutil

repo = 'C:\\Users\\johan\\Bureau\\Data_Lineage'


def main():

################## Génération des fichiers Excel de simulation actuarielle ####

    # Initiate generator objects
    data_gen = DataGenerator()
    excel_gen = ExcelGenerator()

    ## Fichier Pricing

    # Générer les données pour les onglets
    base_params = data_gen.generate_base_params()
    demographic_data = data_gen.generate_demographic_data()
    claim_data = data_gen.generate_claim_data()
    premium_data = data_gen.generate_premium_data()

    # Enregistrer les données dans un fichier Excel
    wb = Workbook()
    ws = wb.active

    # Ajouter les données au fichier Excel
    excel_gen.append_data_to_excel(base_params, "Base_Params", wb)
    excel_gen.append_data_to_excel(demographic_data, "Demographic_Data", wb)
    excel_gen.append_data_to_excel(claim_data, "Claim_Data", wb)
    excel_gen.append_data_to_excel(premium_data, "Premium_Data", wb)

    # Création d'un nouvel onglet pour les calculs actuariels
    ws = wb.create_sheet(title="Actuarial_Calculations")

    # Écrire les formules de calcul avec des liens entre les onglets
    ws.append(['Pure_Premium', 'RBNS', 'Loss_Ratio'])
    ws.append(['=SUM(Claim_Data!D2:D1001)/COUNTA(Premium_Data!C2:C1001)', 
               '=0.1*SUM(Claim_Data!D2:D1001)', 
               '=SUM(Claim_Data!D2:D1001)/SUM(Premium_Data!C2:C1001)'])

    # Ajoutons des onglets supplémentaires avec des formules pour augmenter la complexité
    for i in range(6, 13):  # Créer 7 onglets supplémentaires pour avoir un total de 12 onglets
        ws = wb.create_sheet(title=f"Additional_Calculations_{i}")
        ws.append(['Calculation_1', 'Calculation_2', 'Calculation_3'])
        ws.append([f'=SUM(Base_Params!B2:B1001)*{i}', 
                   f'=SUM(Demographic_Data!B2:B1001)*{i}', 
                   f'=SUM(Premium_Data!C2:C1001)*{i}'])

    # Enregistrer le fichier
    file = 'Pricing.xlsx'
    directory = 'Pricing_directory'
    path_file = excel_gen.generate_xls_and_folder(wb, file, directory)


    ## Fichier Reserving

    # Initiate generator objects
    data_gen = DataGenerator()
    excel_gen = ExcelGenerator()

    # Générer les données pour les onglets
    contract_data = data_gen.generate_contract_data()
    claim_history_data = data_gen.generate_claim_history_data()

    # Générer tous les ensembles de paramètres actuariels à l'avance
    actuarial_params_list = [data_gen.generate_actuarial_params(i) for i in range(8)]

    # Enregistrer les données dans un fichier Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Contract_Data"
    excel_gen.append_data_to_excel(contract_data, "Contract_Data", wb)
    excel_gen.append_data_to_excel(claim_history_data, "Claim_History", wb)
    excel_gen.append_data_to_excel(actuarial_params_list[0], "Actuarial_Params", wb)

    # Création de nouveaux onglets pour les calculs de réserves
    for i in range(1, 8):  # Créer 7 onglets supplémentaires pour avoir un total de 8 onglets
        excel_gen.append_calculation_sheet(actuarial_params_list[i], f"Reserving_Calculations_{i}", wb)

    # Enregistrer le fichier
    file = 'Reserving.xlsx'
    directory = 'Reserving_directory'
    path_file = excel_gen.generate_xls_and_folder(wb, file, directory)


    ## Fichier RiskManagement

    # Initiate generator objects
    data_gen = DataGenerator()
    excel_gen = ExcelGenerator()

    # Generate data for RiskManagement.xlsx
    stress_scenarios = data_gen.generate_stress_scenarios()
    risk_params = data_gen.generate_risk_params()
    risk_exposures = data_gen.generate_risk_exposures()

    # Generate RiskManagement.xlsx file
    wb_risk_management = Workbook()
    excel_gen.append_data_to_excel(stress_scenarios, 'Stress_Scenarios', wb_risk_management)
    excel_gen.append_data_to_excel(risk_params, 'Risk_Params', wb_risk_management)
    excel_gen.append_data_to_excel(risk_exposures, 'Risk_Exposures', wb_risk_management)

    # We also need to add 7 calculation sheets
    for i in range(1, 8):
        risk_exposures = data_gen.generate_risk_exposures(n_rows=1000 + i * 100)
        excel_gen.append_risk_management_calculation_sheet(risk_exposures, f'Risk_Management_Calculations_{i}', wb_risk_management)

    path_risk_management = excel_gen.generate_xls_and_folder(wb_risk_management, "RiskManagement.xlsx", "RiskManagement_directory")


    ## Fichier Solvency

    # Initiate generator objects
    data_gen = DataGenerator()
    excel_gen = ExcelGenerator()

    # Generate asset, liability and regulatory data
    asset_data = data_gen.generate_asset_data()
    liability_data = data_gen.generate_liability_data()
    regulatory_data = data_gen.generate_regulatory_data()

    # Create a new workbook for solvency
    wb_solvency = Workbook()

    # Delete the default sheet
    del wb_solvency['Sheet']

    # Append data to the workbook
    excel_gen.append_data_to_excel(asset_data, "Asset_Data", wb_solvency)
    excel_gen.append_data_to_excel(liability_data, "Liability_Data", wb_solvency)
    excel_gen.append_data_to_excel(regulatory_data, "Regulatory_Data", wb_solvency)

    # Append solvency calculation sheet
    excel_gen.append_solvency_calculation_sheet(asset_data, liability_data, regulatory_data, "Solvency_Calculations", wb_solvency)

    # Generate additional sheets
    for i in range(1, 7):
        # Generate new asset and liability data
        asset_data = data_gen.generate_asset_data(n_rows=2000+i*100)
        liability_data = data_gen.generate_liability_data(n_rows=2000+i*100)

        # Append new data to the workbook
        excel_gen.append_data_to_excel(asset_data, f"Asset_Data_{i}", wb_solvency)
        excel_gen.append_data_to_excel(liability_data, f"Liability_Data_{i}", wb_solvency)

        # Append new solvency calculation sheet
        excel_gen.append_solvency_calculation_sheet(asset_data, liability_data, regulatory_data, f"Solvency_Calculations_{i}", wb_solvency)

    # Generate the Excel file
    path_solvency = excel_gen.generate_xls_and_folder(wb_solvency, "Solvency.xlsx", "Solvency_directory")


    ## Fichier Capital_Modeling

    # Initiate generator objects
    data_gen = DataGenerator()
    excel_gen = ExcelGenerator()

    # Generate data for Capital_Modeling.xlsx
    ids = data_gen.generate_ids(n_rows=2000)
    model_parameters = data_gen.generate_model_parameters()
    model_results = data_gen.generate_model_results(ids)
    monte_carlo_data = data_gen.generate_monte_carlo_data(ids)

    # Generate Capital_Modeling.xlsx file
    wb_capital_modeling = Workbook()

    # Add data to the workbook
    excel_gen.append_data_to_excel(model_parameters, 'Model_Parameters', wb_capital_modeling)
    excel_gen.append_data_to_excel(model_results, 'Model_Results', wb_capital_modeling)
    excel_gen.append_data_to_excel(monte_carlo_data, 'Monte_Carlo_Data', wb_capital_modeling)
    excel_gen.append_capital_modeling_calculations_sheet(ids, 'Capital_Calculations', wb_capital_modeling)

    # Generate and save workbook
    path_capital_modeling = excel_gen.generate_xls_and_folder(wb_capital_modeling, "Capital_Modeling.xlsx", "Capital_Modeling_directory")

    # Reopen and save workbook to add formulas
    excel_gen.reopen_and_save_workbook(path_capital_modeling)


    ## Fichier Reinsurance

    # Initiate generator objects
    data_gen = DataGenerator()
    excel_gen = ExcelGenerator()

    # Generate data for Reinsurance.xlsx
    ids = data_gen.generate_ids()
    reinsurance_contracts = data_gen.generate_reinsurance_contracts()
    claim_history = data_gen.generate_claim_history(reinsurance_contracts['Contract_ID'])
    reinsurance_pricing_params = data_gen.generate_reinsurance_pricing_params()

    # Generate Reinsurance.xlsx file
    wb_reinsurance = Workbook()

    # Add data to the workbook
    excel_gen.append_data_to_excel(reinsurance_contracts, 'Reinsurance_Contracts', wb_reinsurance)
    excel_gen.append_data_to_excel(claim_history, 'Claim_History', wb_reinsurance)
    excel_gen.append_data_to_excel(reinsurance_pricing_params, 'Reinsurance_Pricing_Parameters', wb_reinsurance)
    excel_gen.append_reinsurance_pricing_calculations_sheet('Reinsurance_Pricing_Calculations', wb_reinsurance)

    # Generate and save workbook
    path_reinsurance = excel_gen.generate_xls_and_folder(wb_reinsurance, "Reinsurance.xlsx", "Reinsurance_directory")

    # Reopen and save workbook to add formulas
    excel_gen.reopen_and_save_workbook(path_reinsurance)


    ## Fichier Insurance_Data

    # Initiate generator objects
    data_gen = DataGenerator()
    excel_gen = ExcelGenerator()

    # Générer les données pour les onglets
    base_params = data_gen.generate_base_params()
    demographic_data = data_gen.generate_demographic_data()
    claim_data = data_gen.generate_claim_data()
    premium_data = data_gen.generate_premium_data()
    market_indices = data_gen.generate_market_indices()
    customer_details = data_gen.generate_customer_details()
    policy_details = data_gen.generate_policy_details()

    # Enregistrer les données dans un fichier Excel
    wb = Workbook()

    # Ajouter les données au fichier Excel
    excel_gen.append_data_to_excel(base_params, "Base_Params", wb)
    excel_gen.append_data_to_excel(demographic_data, "Demographic_Data", wb)
    excel_gen.append_data_to_excel(claim_data, "Claim_Data", wb)
    excel_gen.append_data_to_excel(premium_data, "Premium_Data", wb)
    excel_gen.append_data_to_excel(market_indices, "Market_Indices", wb)
    excel_gen.append_data_to_excel(customer_details, "Customer_Details", wb)
    excel_gen.append_data_to_excel(policy_details, "Policy_Details", wb)

    # Append calculation sheets to Excel file
    excel_gen.append_actuarial_calculation_sheet("Actuarial_Calculations", wb)

    # Save the Excel file and the folder
    path_insurance = excel_gen.generate_xls_and_folder(wb, "insurance_data.xlsx", "Insurance_Data_directory")


    ## Autres petits fichiers
    """
    # Fichier Market_Data

    # Initiate generator objects
    data_gen = DataGenerator()
    excel_gen = ExcelGenerator()

    # Générer les données pour le fichier
    indices_data = data_gen.generate_indices_data()
    fund_data = data_gen.generate_fund_data()
    equity_data = data_gen.generate_equity_data()

    # Enregistrer les données dans un fichier Excel
    wb_market_data = Workbook()
    excel_gen.append_data_to_excel(indices_data, "Indices", wb_market_data)
    excel_gen.append_data_to_excel(fund_data, "Funds", wb_market_data)
    excel_gen.append_data_to_excel(equity_data, "Equities", wb_market_data)

    # Enregistrer le fichier
    file = 'Market_Data.xlsx'
    directory = 'Market_Data_directory'
    path_file = excel_gen.generate_xls_and_folder(wb_market_data, file, directory)

    # Fichier Product_Data

    # Initiate generator objects
    data_gen = DataGenerator()
    excel_gen = ExcelGenerator()

    # Générer les données pour le fichier
    product_data = data_gen.generate_product_data()
    product_parameters = data_gen.generate_product_parameters()
    product_pricing = data_gen.generate_product_pricing()

    # Enregistrer les données dans un fichier Excel
    wb_product_data = Workbook()
    excel_gen.append_data_to_excel(product_data, "Products", wb_product_data)
    excel_gen.append_data_to_excel(product_parameters, "Parameters", wb_product_data)
    excel_gen.append_data_to_excel(product_pricing, "Pricing", wb_product_data)

    # Enregistrer le fichier
    file = 'Product_Data.xlsx'
    directory = 'Product_Data_directory'
    path_file = excel_gen.generate_xls_and_folder(wb_product_data, file, directory)

    # Fichier Agent_Data

    # Initiate generator objects
    data_gen = DataGenerator()
    excel_gen = ExcelGenerator()

    # Générer les données pour le fichier
    agent_data = data_gen.generate_agent_data()
    agent_performance = data_gen.generate_agent_performance()

    # Enregistrer les données dans un fichier Excel
    wb_agent_data = Workbook()
    excel_gen.append_data_to_excel(agent_data, "Agents", wb_agent_data)
    excel_gen.append_data_to_excel(agent_performance, "Performance", wb_agent_data)

    # Enregistrer le fichier
    file = 'Agent_Data.xlsx'
    directory = 'Agent_Data_directory'
    path_file = excel_gen.generate_xls_and_folder(wb_agent_data, file, directory)
    """

    data_generator = DataGenerator()
    excel_generator = ExcelGenerator()

    datasets = {
        "Market_Data": data_generator.generate_market_indices(),
        "Customer_Data": data_generator.generate_customer_details(),
        "Policy_Data": data_generator.generate_policy_details(),
        "Claim_Data": data_generator.generate_claim_details(),
        "Agent_Data": data_generator.generate_agent_details(),
        "Product_Data": data_generator.generate_product_details(),
    }

    for name, data in datasets.items():
        wb = Workbook()
        ws = wb.active
        ws.title = name
        for r in dataframe_to_rows(data, index=False, header=True):
            ws.append(r)
        file = f'{name}.xlsx'
        directory = f'{name}_directory'
        path_file = excel_generator.generate_xls_and_folder(wb, file, directory)


    ## Structuration en arborescence des fichiers Excel

    # Create folder structure
    os.chdir(repo)
    root_folder = 'Insurance_Company' # os.getcwd()
    folders = {
        'Actuarial': ['Pricing_and_Modelling'],
        'Customer_Relations': [],
        'Claims_and_Policy': [],
        'Finance': []
    }
    for parent_folder, subfolders in folders.items():
        parent_path = os.path.join(root_folder, parent_folder)
        os.makedirs(parent_path, exist_ok=True)
        for subfolder in subfolders:
            os.makedirs(os.path.join(parent_path, subfolder), exist_ok=True)

    # Move files to appropriate folders
    file_folder_mapping = {
        'Agent_Data.xlsx': os.path.join(root_folder, 'Customer_Relations'),
        'Capital_Modeling.xlsx': os.path.join(root_folder, 'Actuarial', 'Pricing_and_Modelling'),
        'Claim_Data.xlsx': os.path.join(root_folder, 'Claims_and_Policy'),
        'Customer_Data.xlsx': os.path.join(root_folder, 'Customer_Relations'),
        'Market_Data.xlsx': os.path.join(root_folder, 'Finance'),
        'Policy_Data.xlsx': os.path.join(root_folder, 'Claims_and_Policy'),
        'Pricing.xlsx': os.path.join(root_folder, 'Actuarial', 'Pricing_and_Modelling'),
        'Product_Data.xlsx': os.path.join(root_folder, 'Customer_Relations'),
        'Reinsurance.xlsx': os.path.join(root_folder, 'Finance'),
        'Reserving.xlsx': os.path.join(root_folder, 'Claims_and_Policy'),
        'RiskManagement.xlsx': os.path.join(root_folder, 'Actuarial'),
        'Solvency.xlsx': os.path.join(root_folder, 'Actuarial')
    }

    for file, folder in file_folder_mapping.items():
        old_path = os.path.join(file.rsplit('.')[0] + '_directory', file)
        new_path = os.path.join(folder, file)
        shutil.move(old_path, new_path)

    # Supprimer les anciens dossiers
    for old_folder in os.listdir():
        if old_folder.endswith('_directory'):
            shutil.rmtree(old_folder)


    ## Ajouter les clés étrangères

    from generation.excel_generator import ExcelModifier

    files_to_update = {
        'Actuarial/Pricing_and_Modelling/Pricing.xlsx': ['Product_ID', 'Customer_ID'],
        'Actuarial/Pricing_and_Modelling/Capital_Modeling.xlsx': ['Scenario_ID'],
        'Actuarial/RiskManagement.xlsx': ['Scenario_ID'],
        'Actuarial/Solvency.xlsx': ['Asset_ID', 'Liability_ID'],
        'Customer_Relations/Agent_Data.xlsx': ['Customer_ID'],
        'Customer_Relations/Customer_Data.xlsx': ['Product_ID'],
        'Claims_and_Policy/Claim_Data.xlsx': ['Policy_ID'],
        'Claims_and_Policy/Policy_Data.xlsx': ['Customer_ID', 'Product_ID'],
        'Claims_and_Policy/Reserving.xlsx': ['Contract_ID', 'Claim_ID'],
        'Finance/Market_Data.xlsx': ['Index_ID'],
        'Finance/Reinsurance.xlsx': ['Contract_ID', 'Claim_ID'],
    }

    data_dir = 'Insurance_Company'
    excel_modifier = ExcelModifier(data_dir)

    for file, columns in files_to_update.items():
        excel_modifier.add_foreign_keys(file, columns)



################## Génération du graphe à partir des fichiers Excel ###########

    # Création de l'instance de GraphGenerator
    graph_gen = GraphGenerator(data_dir=repo)

    # Génération du graphe à partir des fichiers
    graph_gen.walk_files_and_generate_graph()

    # Ajout des clés étrangères au graphe
    graph_gen.add_foreign_keys()

    # Enregistrement du graphe dans un fichier
    graph_gen.save_graph("lineage_graph.gexf")

    # Création des sous-graphes et visualisation
    graph_gen.create_subgraphs_and_visualize()


###############################################################################
        
if __name__ == "__main__":
    main()