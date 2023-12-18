import os
import shutil
import random
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

class ExcelGenerator:
    def __init__(self, data_dir="C:/Users/johan/Bureau/Data_Lineage"):
        self.data_dir = data_dir

    def append_data_to_excel(self, df, sheet_name, workbook):
        if not isinstance(df, pd.DataFrame):
            raise TypeError('df doit être un DataFrame pandas.')
        if df.empty:
            raise ValueError('df ne peut pas être vide.')
        if not isinstance(sheet_name, str):
            raise TypeError('sheet_name doit être une chaîne de caractères.')
        if not sheet_name:
            raise ValueError('sheet_name ne peut pas être vide.')
        if not isinstance(workbook, openpyxl.workbook.workbook.Workbook):
            raise TypeError('workbook doit être un objet Workbook de openpyxl.')

        ws = workbook.create_sheet(title=sheet_name)
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

    # def add_foreign_keys(self, workbook, column_names, id_list):
    #     for sheet_name in workbook.sheetnames:
    #         ws = workbook[sheet_name]
    #         if ws.max_row > 1:  # Check if sheet has more than one row
    #             for column_name in column_names:
    #                 col_idx = ws.max_column + 1
    #                 ws.cell(row=1, column=col_idx, value=column_name)
    #                 for row in range(2, ws.max_row + 1):
    #                     ws.cell(row=row, column=col_idx, value=random.choice(id_list))  # Use existing IDs
    #     return workbook
        
    def generate_xls_and_folder(self, wb, file, directory):
        """Génère un dossier avec le xls à l'intérieur"""
        path_dir = os.path.join(self.data_dir, directory)
        if os.path.exists(path_dir) and os.path.isdir(path_dir):
            shutil.rmtree(path_dir)
        os.mkdir(path_dir)
        path_file = os.path.join(path_dir, file)
        wb.save(path_file)
        return path_file

    def append_calculation_sheet(self, actuarial_params, sheet_name, workbook):
        """Ajoute un nouvel onglet avec des formules de calcul à un classeur Excel."""
        ws = workbook.create_sheet(title=sheet_name)
        ws.append(['RNR', 'PRC', 'Sinistralité', 'Loss_Development', 'Expense_Ratio'])
        ws.append(['=0.1*SUM(Claim_History!E2:E1001)', 
                   '=SUM(Contract_Data!C2:C1001) - (SUM(Contract_Data!C2:C1001) / (1 + ' + str(actuarial_params['Parameter_Value'][0]) + '))', 
                   '=SUM(Claim_History!E2:E1001)/SUM(Contract_Data!C2:C1001)',
                   '=SUM(Claim_History!E2:E1001)*' + str(actuarial_params['Parameter_Value'][1]),
                   '=SUM(Claim_History!E2:E1001)*' + str(actuarial_params['Parameter_Value'][2])])

    def append_risk_management_calculation_sheet(self, risk_exposures, sheet_name, workbook):
        """Ajoute un nouvel onglet avec des formules de calcul à un classeur Excel pour le RiskManagement."""
        ws = workbook.create_sheet(title=sheet_name)
        
        # Add column names first
        ws.append(['Asset_ID', 'VaR', 'TVar', 'Expected_Shortfall', 'Scenario_VaR', 'Scenario_TVaR'])
        
        # Add data with formulas
        for j in range(2, len(risk_exposures)+2):
            ws.append([f'=Risk_Exposures!A{j}',
                       f'=PERCENTILE(Stress_Scenarios!C2:C{len(risk_exposures)+1}, 0.995) * AVERAGE(Risk_Exposures!B2:B{len(risk_exposures)+1})',
                       f'=AVERAGEIF(Stress_Scenarios!C2:C{len(risk_exposures)+1}, ">="&PERCENTILE(Stress_Scenarios!C2:C{len(risk_exposures)+1}, 0.995)) * AVERAGE(Risk_Exposures!B2:B{len(risk_exposures)+1})',
                       f'=PERCENTILE(Stress_Scenarios!C2:C{len(risk_exposures)+1}, 0.975) * AVERAGE(Risk_Exposures!B2:B{len(risk_exposures)+1})',
                       f'=PERCENTILE(Stress_Scenarios!C2:C{len(risk_exposures)+1}, 0.995) * VLOOKUP(A{j}, Risk_Exposures!A2:B{len(risk_exposures)+1}, 2, FALSE)',
                       f'=AVERAGEIF(Stress_Scenarios!C2:C{len(risk_exposures)+1}, ">="&PERCENTILE(Stress_Scenarios!C2:C{len(risk_exposures)+1}, 0.995)) * VLOOKUP(A{j}, Risk_Exposures!A2:B{len(risk_exposures)+1}, 2, FALSE)'])

    def append_solvency_calculation_sheet(self, asset_data, liability_data, regulatory_data, sheet_name, workbook):
        """Ajoute un nouvel onglet avec des formules de calcul à un classeur Excel pour le Solvency."""
        ws = workbook.create_sheet(title=sheet_name)
        
        # Add formulas
        ws.append(['Available_Capital', 'Required_Capital', 'Solvency_Ratio'])
        ws.append([f'=SUM(Asset_Data!B2:B{len(asset_data)+1}) - SUM(Liability_Data!B2:B{len(liability_data)+1})', 
                   f'=MAX(Regulatory_Data!B2, SUM(Liability_Data!B2:B{len(liability_data)+1}) * Regulatory_Data!B3)',
                   f'={ws.title}!A2 / {ws.title}!B2'])

    def append_capital_modeling_calculations_sheet(self, ids, sheetname, workbook):
        ws = workbook.create_sheet(title=sheetname)
        ws.append(['ID', 'Economic_Capital', 'VaR', 'TVar', 'Scenario_Result', 'Simulation_Result'])
        
        for i in range(2, len(ids)+2):
            ws.append([ids[i-2],
                       '=SUM(Model_Results!B2:B2001)', 
                       '=PERCENTILE(Model_Results!B2:B2001, 0.995)', 
                       '=AVERAGEIF(Model_Results!B2:B2001, ">="&' + sheetname + '!B' + str(i) + ')',
                       '=VLOOKUP(A' + str(i) + ', Model_Results!A2:B2001, 2, FALSE)',
                       '=VLOOKUP(A' + str(i) + ', Monte_Carlo_Data!A2:B2001, 2, FALSE)'])

    def reopen_and_save_workbook(self, filepath):
        from openpyxl import load_workbook
        wb = load_workbook(filepath)
        wb.save(filepath)

    def append_reinsurance_pricing_calculations_sheet(self, sheet_name, workbook):
        ws = workbook.create_sheet(title=sheet_name)

        ws.append(['Net_Premium', 'Gross_Premium', 'Commercial_Premium'])
        ws.append(['=SUM(Claim_History!C2:C1001)/COUNTA(Reinsurance_Contracts!D2:D1001)', 
                   '=Reinsurance_Pricing_Calculations!A2/(1-Reinsurance_Pricing_Parameters!B3)', 
                   '=Reinsurance_Pricing_Calculations!B2*(1+Reinsurance_Pricing_Parameters!B2)'])

        return workbook

    def append_market_indices_to_excel(self, df, sheet_name, workbook):
        ws = workbook.create_sheet(title=sheet_name)
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

    def append_customer_details_to_excel(self, df, sheet_name, workbook):
        ws = workbook.create_sheet(title=sheet_name)
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

    def append_policy_details_to_excel(self, df, sheet_name, workbook):
        ws = workbook.create_sheet(title=sheet_name)
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

    def append_actuarial_calculation_sheet(self, sheet_name, workbook):
        """Ajoute un nouvel onglet avec des formules de calcul à un classeur Excel."""
        ws = workbook.create_sheet(title=sheet_name)
        ws.append(['Pure_Premium', 'RBNS', 'Loss_Ratio'])
        ws.append(['=SUM(Claim_Data!D2:D1001)/COUNTA(Premium_Data!C2:C1001)', 
                   '=0.1*SUM(Claim_Data!D2:D1001)', 
                   '=SUM(Claim_Data!D2:D1001)/SUM(Premium_Data!C2:C1001)'])

class ExcelModifier:
    def __init__(self, base_dir):
        self.base_dir = base_dir

    def add_foreign_keys(self, file_path, column_names):
        wb = openpyxl.load_workbook(filename=os.path.join(self.base_dir, file_path))
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if ws.max_row > 1:  # Check if sheet has more than one row
                for column_name in column_names:
                    col_idx = ws.max_column + 1
                    ws.cell(row=1, column=col_idx, value=column_name)
                    for row in range(2, ws.max_row + 1):
                        ws.cell(row=row, column=col_idx, value=str(random.randint(1, 100)))
        wb.save(os.path.join(self.base_dir, file_path))