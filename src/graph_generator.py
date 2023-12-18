import networkx as nx
import openpyxl
import os
import re
from collections import defaultdict
import matplotlib.pyplot as plt

class GraphGenerator:
    def __init__(self, data_dir):
        self.data_dir = data_dir
        self.G = nx.DiGraph()
        self.foreign_keys = defaultdict(list)

    def extract_column_references(self, formula, column_headers):
        """Extract the column references from a cell formula and translate them into column header names"""
        cell_references = re.findall(r'\b[A-Z]{1,3}\d+\b', formula)
        column_references = {ref[:re.search("\d", ref).start()] for ref in cell_references}

        # Translate column references to column headers
        header_references = [header for header in column_headers if header in column_references]
        
        # Extract sheet references
        sheet_references = re.findall(r"'([A-Za-z_ ]+)'!", formula)
        
        return header_references, sheet_references

    def walk_files_and_generate_graph(self):
        for root, dirs, files in os.walk(self.data_dir):
            for file in files:
                if file.endswith(".xlsx"):
                    wb = openpyxl.load_workbook(os.path.join(root, file), data_only=False)
                    for sheet_name in wb.sheetnames:
                        sheet = wb[sheet_name]
                        column_headers = {cell.column_letter: cell.value for cell in sheet[1]}

                        for header, column in list(column_headers.items())[1:]:
                            if str(column).endswith('_ID'):
                                self.foreign_keys[column].append(f"{os.path.splitext(file)[0]}\n{sheet_name}\n{column}")

                        for row in sheet.iter_rows(min_row=2):
                            for cell in row:
                                if cell.data_type == "f":  # la cellule contient une formule
                                    source = f"{os.path.splitext(file)[0]}\n{sheet_name}\n{column_headers[cell.column_letter]}"
                                    self.G.add_node(source)
                                    references = self.extract_column_references(cell.value, column_headers.values())
                                    for ref in references:
                                        if ref in column_headers.values(): # checking in values instead of keys
                                            target = f"{os.path.splitext(file)[0]}\n{sheet_name}\n{ref}" # ref is already column name
                                            self.G.add_node(target)
                                            self.G.add_edge(source, target)

    def add_foreign_keys(self):
        for key, files in self.foreign_keys.items():
            if len(files) > 1:  # make sure there is more than one node for a given foreign key
                for i in range(len(files) - 1):
                    for j in range(i + 1, len(files)):
                        self.G.add_edge(files[i], files[j], color='#ff0000')  # red for foreign key edges

    def save_graph(self, file_name):
        nx.write_gexf(self.G, file_name)

    def create_subgraphs_and_visualize(self):
        # Création du graphe sans les clés étrangères
        G_no_fk = nx.DiGraph()

        # Parcourir les dossiers et fichiers
        for root, dirs, files in os.walk(self.data_dir):
            for file in files:
                if file.endswith(".xlsx"):
                    wb = openpyxl.load_workbook(os.path.join(root, file), data_only=False)
                    for sheet_name in wb.sheetnames:
                        sheet = wb[sheet_name]
                        column_headers = {cell.column_letter: cell.value for cell in sheet[1]}

                        for row in sheet.iter_rows(min_row=2):
                            if any(cell.data_type == "f" for cell in row):  # the row contains at least one formula
                                for cell in row:
                                    if cell.data_type == "f":  # the cell contains a formula
                                        source = f"{os.path.splitext(file)[0]}\n{sheet_name}\n{column_headers[cell.column_letter]}"
                                        G_no_fk.add_node(source)
                                        header_references, sheet_references = self.extract_column_references(cell.value, column_headers)
                                        
                                        # Check if there are any sheet references, then adjust target node to reflect that.
                                        if sheet_references:
                                            for ref_sheet in sheet_references:
                                                for ref in header_references:
                                                    target = f"{os.path.splitext(file)[0]}\n{ref_sheet}\n{ref}"
                                                    G_no_fk.add_node(target)
                                                    G_no_fk.add_edge(source, target)
                                        else:
                                            for ref in header_references:
                                                target = f"{os.path.splitext(file)[0]}\n{sheet_name}\n{ref}"
                                                G_no_fk.add_node(target)
                                                G_no_fk.add_edge(source, target)

        # Convertir le graphe dirigé en graphe non dirigé pour obtenir les sous-graphes
        UG_no_fk = G_no_fk.to_undirected()

        # Obtenir les sous-graphes
        subgraphs_no_fk = [G_no_fk.subgraph(c) for c in nx.connected_components(UG_no_fk)]

        # Filtrer les sous-graphes pour ne conserver que ceux qui contiennent des noeuds de "Capital_Modeling.xlsx"
        filtered_subgraphs_no_fk = [sg for sg in subgraphs_no_fk if any("Capital_Modeling" in node for node in sg.nodes)]

        for i in range(len(filtered_subgraphs_no_fk)):
            plt.figure(figsize=(6, 6), dpi=600)
            nx.draw(filtered_subgraphs_no_fk[i], with_labels=True, arrows=False, font_size=4)
            plt.show()
